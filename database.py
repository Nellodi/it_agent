# Файл: it_ecosystem_bot/database.py
import sqlite3
import asyncio
import logging
import time
import os
from typing import Dict, Any, List, Tuple, Optional
import re

logger = logging.getLogger(__name__)

DB_PATH = 'it_ecosystem.db'


# --- ФУНКЦИИ МИГРАЦИИ И ДОПОЛНЕНИЯ ТАБЛИЦ ---
def _ensure_workplaces_columns_and_tables():
    """Проверяет и добавляет необходимые колонки в существующие таблицы (миграция)."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # --- Миграция workplaces ---
    cursor.execute("PRAGMA table_info(workplaces)")
    cols = {r[1] for r in cursor.fetchall()}
    if 'floor' not in cols:
        try:
            cursor.execute("ALTER TABLE workplaces ADD COLUMN floor INTEGER")
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка добавления колонки floor: {e}")
    if 'primary_pc' not in cols:
        try:
            cursor.execute("ALTER TABLE workplaces ADD COLUMN primary_pc TEXT")
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка добавления колонки primary_pc: {e}")
    if 'peripherals' not in cols:
        try:
            cursor.execute("ALTER TABLE workplaces ADD COLUMN peripherals TEXT")
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка добавления колонки peripherals: {e}")

    # --- Миграция authorized_users (email) ---
    cursor.execute("PRAGMA table_info(authorized_users)")
    auth_cols = {r[1] for r in cursor.fetchall()}
    if 'email' not in auth_cols:
        try:
            cursor.execute("ALTER TABLE authorized_users ADD COLUMN email TEXT")
        except sqlite3.Error:
            pass

    # --- Миграция tickets (pc_name, floor) ---
    cursor.execute("PRAGMA table_info(tickets)")
    ticket_cols = {r[1] for r in cursor.fetchall()}
    if 'pc_name' not in ticket_cols:
        try:
            cursor.execute("ALTER TABLE tickets ADD COLUMN pc_name TEXT DEFAULT NULL")
        except sqlite3.Error:
            pass
    if 'floor' not in ticket_cols:
        try:
            cursor.execute("ALTER TABLE tickets ADD COLUMN floor INTEGER DEFAULT NULL")
        except sqlite3.Error:
            pass

    # --- Создание дополнительных таблиц, если отсутствуют ---
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS user_credentials (
            id INTEGER PRIMARY KEY AUTOINCREMENT, telegram_id INTEGER NOT NULL, service TEXT NOT NULL, login TEXT, password TEXT, url TEXT, note TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, FOREIGN KEY (telegram_id) REFERENCES authorized_users (telegram_id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS peripherals_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT, workplace_id INTEGER, pc_hostname TEXT, device_type TEXT, action TEXT, details TEXT, reported_by INTEGER,
            event_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP, FOREIGN KEY (workplace_id) REFERENCES workplaces (id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ticket_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT, ticket_id INTEGER NOT NULL, file_id TEXT NOT NULL, file_type TEXT, file_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, FOREIGN KEY (ticket_id) REFERENCES tickets (id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ticket_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT, ticket_id INTEGER NOT NULL, old_status TEXT, new_status TEXT,
            changed_by INTEGER, changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, comment TEXT,
            FOREIGN KEY (ticket_id) REFERENCES tickets (id), FOREIGN KEY (changed_by) REFERENCES authorized_users (telegram_id)
        )
    """)

    conn.commit()
    conn.close()

def _seed_default_workplaces():
    """Автосоздание рабочих мест по умолчанию, если их нет."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for sql in [
        'ALTER TABLE workplaces ADD COLUMN floor INTEGER',
        'ALTER TABLE workplaces ADD COLUMN primary_pc TEXT',
        'ALTER TABLE workplaces ADD COLUMN peripherals TEXT',
    ]:
        try:
            c.execute(sql)
        except Exception:
            pass

    def upsert(number: str, floor: int):
        c.execute('SELECT id FROM workplaces WHERE number=?', (number,))
        row = c.fetchone()
        if row:
            c.execute('UPDATE workplaces SET floor=?, primary_pc=? WHERE id=?', (floor, number, row[0]))
        else:
            c.execute('INSERT INTO workplaces (number, department, location, floor, primary_pc) VALUES (?, ?, ?, ?, ?)',
                      (number, None, None, floor, number),)

    for num in range(2001, 2093):
        upsert(f'TSS-WS-{num}', 2)
    for num in range(1, 73):
        upsert(f'TSS-WS-{num:03d}', 4)
    for num in range(5001, 5063):
        upsert(f'TSS-WS-{num}', 5)

    conn.commit()
    conn.close()


def import_users_from_excel(file_path: str = "users.xlsx"):
    """Импортирует пользователей из Excel в auth_directory (замена Excel-парсера)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl не установлен, импорт из Excel невозможен.")
        return

    if not os.path.exists(file_path):
        logger.warning(f"Файл {file_path} не найден, пропускаем импорт.")
        return

    wb = load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    def col(name: str) -> int | None:
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx_name = col("Name")
    idx_dept = col("Department")
    idx_pos = col("Position")
    idx_login = col("Login")
    idx_pass = col("Password")
    idx_email = col("Email Address")
    idx_status = col("Status")

    if idx_login is None or idx_pass is None:
        logger.error("В Excel нет столбцов Login/Password, импорт невозможен.")
        return

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        login = (row[idx_login] or "").strip().lower()
        password = (row[idx_pass] or "").strip()
        if not login or not password:
            continue

        status = (row[idx_status] or "").lower() if idx_status is not None else ""
        if "terminated" in status or "❌" in status:
            continue  # пропускаем уволенных

        full_name = row[idx_name] if idx_name is not None else None
        dept = row[idx_dept] if idx_dept is not None else None
        pos = row[idx_pos] if idx_pos is not None else None
        email = row[idx_email] if idx_email is not None else None

        try:
            c.execute("""
                INSERT OR REPLACE INTO auth_directory (login, password, full_name, department, position, role, email)
                VALUES (?, ?, ?, ?, ?, COALESCE((SELECT role FROM auth_directory WHERE login=?), 'user'), ?)
            """, (login, password, full_name, dept, pos, login, email))
            count += 1
        except sqlite3.Error as e:
            logger.error(f"DB: ошибка импорта пользователя {login}: {e}")

    conn.commit()
    conn.close()
    logger.info(f"Импортировано пользователей в auth_directory: {count}")


async def get_auth_directory_user(login: str) -> Optional[Dict[str, str]]:
    """Получить запись пользователя из справочника auth_directory по логину."""

    def _get():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT login, password, full_name, department, position, role, email
            FROM auth_directory
            WHERE login = ?
        """, (login.lower(),))
        row = cursor.fetchone()
        conn.close()
        if not row:
            return None
        return {
            "login": row[0],
            "password": row[1],
            "full_name": row[2],
            "department": row[3],
            "position": row[4],
            "role": row[5],
            "email": row[6],
        }

    return await asyncio.to_thread(_get)



async def init_db():
    """Инициализирует базу данных и создает все необходимые таблицы."""

    def create_tables():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # 1. Таблица: authorized_users
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS authorized_users (
                telegram_id INTEGER PRIMARY KEY, login TEXT UNIQUE, full_name TEXT, department TEXT, position TEXT,
                role TEXT DEFAULT 'user', email TEXT, authorized_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # 2. Таблица: tickets
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS tickets (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, ticket_number TEXT UNIQUE, title TEXT, description TEXT,
                status TEXT DEFAULT 'open', priority TEXT DEFAULT 'medium', category TEXT, admin_id INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, closed_at TIMESTAMP, pc_name TEXT, floor INTEGER,
                FOREIGN KEY (user_id) REFERENCES authorized_users (telegram_id)
            )
        """)

        # 3. Таблица: sys_admins (Для хранения рейтинга администраторов)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS sys_admins (
                telegram_id INTEGER PRIMARY KEY, full_name TEXT, total_rating REAL DEFAULT 0.0, rating_count INTEGER DEFAULT 0, 
                FOREIGN KEY (telegram_id) REFERENCES authorized_users (telegram_id)
            )
        """)

        # 4. Таблица: ticket_history
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ticket_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT, ticket_id INTEGER NOT NULL, old_status TEXT, new_status TEXT,
                changed_by INTEGER, changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, comment TEXT,
                FOREIGN KEY (ticket_id) REFERENCES tickets (id), FOREIGN KEY (changed_by) REFERENCES authorized_users (telegram_id)
            )
        """)

        # 5. Таблица: workplaces (Рабочие места)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS workplaces (
                id INTEGER PRIMARY KEY AUTOINCREMENT, number TEXT UNIQUE NOT NULL, department TEXT, location TEXT, 
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # 6. Таблица: equipment (Оборудование)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS equipment (
                id INTEGER PRIMARY KEY AUTOINCREMENT, inv_number TEXT UNIQUE NOT NULL, model TEXT, serial TEXT, category TEXT,
                status TEXT DEFAULT 'available', user_id INTEGER, workplace_id INTEGER, assigned_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES authorized_users (telegram_id),
                FOREIGN KEY (workplace_id) REFERENCES workplaces (id)
            )
        """)

        # 7. Таблица: equipment_history
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS equipment_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT, equipment_id INTEGER NOT NULL, from_user_id INTEGER, to_user_id INTEGER,
                assigned_by INTEGER, assigned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, reason TEXT,
                FOREIGN KEY (equipment_id) REFERENCES equipment (id), FOREIGN KEY (from_user_id) REFERENCES authorized_users (telegram_id),
                FOREIGN KEY (to_user_id) REFERENCES authorized_users (telegram_id), FOREIGN KEY (assigned_by) REFERENCES authorized_users (telegram_id)
            )
        """)

        # 8. Таблица: licenses
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS licenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, license_number TEXT UNIQUE, product TEXT, version TEXT,
                expiry_date DATE, status TEXT DEFAULT 'active', cost REAL, assigned_to INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (assigned_to) REFERENCES authorized_users (telegram_id)
            )
        """)

        # 9. Таблица: faq_materials (КРИТИЧЕСКИ ВАЖНО)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS faq_materials (
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                title TEXT NOT NULL, 
                description TEXT,
                file_id TEXT,
                file_type TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # 10. Таблица: auth_directory (справочник логинов/паролей)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS auth_directory (
                login TEXT PRIMARY KEY,
                password TEXT NOT NULL,
                full_name TEXT,
                department TEXT,
                position TEXT,
                role TEXT DEFAULT 'user',
                email TEXT
            )
        """)

        conn.commit()
        conn.close()

    await asyncio.to_thread(create_tables)
    await asyncio.to_thread(_ensure_workplaces_columns_and_tables)
    await asyncio.to_thread(_seed_default_workplaces)

# --- ОСНОВНЫЕ ФУНКЦИИ ПОЛЬЗОВАТЕЛЯ И АВТОРИЗАЦИИ ---

async def save_authorized_user(telegram_id: int, user_data: Dict[str, str]):
    """Сохраняет нового авторизованного пользователя в authorized_users."""

    def insert_user():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        login = user_data['login']
        full_name = user_data['full_name']
        department = user_data['department']
        position = user_data['position']
        role = user_data.get('role', 'user')
        email = user_data.get('email', '')
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO authorized_users 
                (telegram_id, login, full_name, department, position, role, email)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (telegram_id, login, full_name, department, position, role, email))
            return True
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка сохранения пользователя {telegram_id}: {e}")
            return False
        finally:
            conn.commit()
            conn.close()

    return await asyncio.to_thread(insert_user)


async def get_user_role(telegram_id: int) -> str | None:
    """Получает роль пользователя."""

    def fetch_role():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT role FROM authorized_users WHERE telegram_id = ?", (telegram_id,))
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else None

    return await asyncio.to_thread(fetch_role)


async def get_full_user_profile(telegram_id: int) -> Dict[str, str] | None:
    """Получает полные данные профиля для отображения."""

    def fetch_profile():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT full_name, login, department, position, role, email, authorized_at
            FROM authorized_users WHERE telegram_id = ?
        """, (telegram_id,))
        result = cursor.fetchone()
        conn.close()
        if result:
            return {
                'full_name': result[0],
                'login': result[1],
                'department': result[2],
                'position': result[3],
                'role': result[4],
                'email': result[5] if result[5] else 'Не указана',
                'authorized_at': result[6].split(' ')[0]
            }
        return None

    return await asyncio.to_thread(fetch_profile)


async def remove_authorized_user(telegram_id: int) -> bool:
    """Удаляет пользователя из таблицы authorized_users (деавторизация)."""

    def delete_user():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM authorized_users WHERE telegram_id = ?", (telegram_id,))
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка удаления пользователя {telegram_id}: {e}")
            return False
        finally:
            conn.commit()
            conn.close()
            return True

    return await asyncio.to_thread(delete_user)


# --- ФУНКЦИИ УПРАВЛЕНИЯ ЗАЯВКАМИ ---

async def save_new_ticket(user_id: int, data: Dict[str, Any]) -> Tuple[int, str]:
    """��������� ����� ������ � ������� (id, �����)."""

    def insert_ticket():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        date_part = time.strftime("%y%m%d", time.localtime())
        cursor.execute(f"SELECT MAX(ticket_number) FROM tickets WHERE ticket_number LIKE 'TK{date_part}%'" )
        last_num = cursor.fetchone()[0]

        sequence = int(last_num[-4:]) + 1 if last_num else 1
        ticket_number = f"TK{date_part}{sequence:04d}"

        columns = ["user_id", "ticket_number", "title", "description", "category", "priority"]
        values = [
            user_id,
            ticket_number,
            data.get('title'),
            data.get('description'),
            data.get('category'),
            data.get('priority', 'medium')
        ]

        if 'floor' in data:
            columns.append("floor")
            values.append(data.get('floor'))
        if 'workplace' in data:
            columns.append("pc_name")
            values.append(data.get('workplace'))

        cols_sql = ", ".join(columns)
        placeholders = ", ".join(["?"] * len(values))
        cursor.execute(f"INSERT INTO tickets ({cols_sql}) VALUES ({placeholders})", values)

        ticket_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return ticket_id, ticket_number

    return await asyncio.to_thread(insert_ticket)


async def get_admin_telegram_ids() -> list[int]:
    """Получает все Telegram ID пользователей с ролью 'admin'."""

    def fetch_admin_ids():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT telegram_id FROM authorized_users WHERE role = 'admin'")
        admin_ids = [row[0] for row in cursor.fetchall()]
        conn.close()
        return admin_ids

    return await asyncio.to_thread(fetch_admin_ids)


async def get_all_tickets(status: str = None, department: str = None, priority: str = None) -> List[Dict]:
    """Получает список всех заявок с опциональной фильтрацией."""

    def fetch_tickets():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        query = """
            SELECT t.id, t.ticket_number, t.title, t.status, t.priority, t.category, 
                   t.user_id, t.admin_id, t.created_at, u.full_name, u.department
            FROM tickets t
            LEFT JOIN authorized_users u ON t.user_id = u.telegram_id
            WHERE 1=1
        """
        params = []
        if status: query += " AND t.status = ?"; params.append(status)
        if priority: query += " AND t.priority = ?"; params.append(priority)
        if department: query += " AND u.department = ?"; params.append(department)

        query += " ORDER BY t.created_at DESC"

        cursor.execute(query, params)
        tickets = []
        for row in cursor.fetchall():
            tickets.append({
                'id': row[0], 'number': row[1], 'title': row[2], 'status': row[3], 'priority': row[4],
                'category': row[5], 'user_id': row[6], 'admin_id': row[7], 'created_at': row[8],
                'user_name': row[9], 'department': row[10]
            })
        conn.close()
        return tickets

    return await asyncio.to_thread(fetch_tickets)


async def get_ticket_by_id(ticket_id: int) -> Dict | None:
    """Возвращает подробную информацию по тикету (по id)."""

    def _get():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT t.*, u.full_name as user_name, a.full_name as admin_name
            FROM tickets t
            JOIN authorized_users u ON t.user_id = u.telegram_id
            LEFT JOIN sys_admins a ON t.admin_id = a.telegram_id
            WHERE t.id = ?
        """, (ticket_id,))

        row = cursor.fetchone()
        conn.close()
        return row  # Возвращает tuple

    return await asyncio.to_thread(_get)


async def assign_ticket_to_admin(ticket_id: int, admin_id: int, old_status: str = 'open') -> bool:
    """Назначает заявку на администратора и устанавливает статус 'in_progress'."""

    def assign_ticket():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("""
                UPDATE tickets SET admin_id = ?, status = 'in_progress' WHERE id = ? AND status = ?
            """, (admin_id, ticket_id, old_status))

            if cursor.rowcount > 0:
                old_status_for_history = 'open'
                cursor.execute("""
                    INSERT INTO ticket_history (ticket_id, old_status, new_status, changed_by, comment)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                ticket_id, old_status_for_history, 'in_progress', admin_id, 'Назначен администратор и начата работа'))

            return cursor.rowcount > 0
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка назначения тикета {ticket_id} администратору {admin_id}: {e}")
            return False
        finally:
            conn.commit()
            conn.close()

    return await asyncio.to_thread(assign_ticket)


async def update_ticket_status(ticket_id: int, new_status: str, admin_id: int, comment: str = None) -> bool:
    """Обновляет статус заявки и создает запись в истории."""

    def update_status():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # 1. Получаем текущий статус для истории
        cursor.execute("SELECT status FROM tickets WHERE id = ?", (ticket_id,))
        old_status = cursor.fetchone()

        if not old_status: return False
        old_status = old_status[0]

        try:
            cursor.execute("""
                UPDATE tickets SET status = ?, admin_id = ? WHERE id = ?
            """, (new_status, admin_id, ticket_id))

            # 2. Создаем запись в истории
            cursor.execute("""
                INSERT INTO ticket_history (ticket_id, old_status, new_status, changed_by, comment)
                VALUES (?, ?, ?, ?, ?)
            """, (ticket_id, old_status, new_status, admin_id, comment))

            return cursor.rowcount > 0
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка обновления статуса тикета {ticket_id}: {e}")
            return False
        finally:
            conn.commit()
            conn.close()

    return await asyncio.to_thread(update_status)


async def get_ticket_history(ticket_id: int) -> List[Dict]:
    """Получает историю изменений статусов для заявки."""

    def fetch_history():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                th.old_status, th.new_status, th.changed_at, th.comment, u.full_name
            FROM ticket_history th
            LEFT JOIN authorized_users u ON th.changed_by = u.telegram_id
            WHERE th.ticket_id = ?
            ORDER BY th.changed_at ASC
        """, (ticket_id,))

        history = []
        for row in cursor.fetchall():
            history.append({
                'old_status': row[0], 'new_status': row[1], 'changed_at': row[2], 'comment': row[3],
                'changed_by_name': row[4] if row[4] else 'Система'
            })
        conn.close()
        return history

    return await asyncio.to_thread(fetch_history)


async def get_user_tickets(user_id: int) -> List[Dict]:
    """Получает список заявок пользователя."""

    def fetch_tickets():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, ticket_number, title, status, created_at, category, priority
            FROM tickets WHERE user_id = ? 
            ORDER BY created_at DESC
        """, (user_id,))

        rows = cursor.fetchall()
        conn.close()

        tickets = []
        for r in rows:
            tickets.append({
                'id': r[0], 'number': r[1], 'title': r[2], 'status': r[3], 'created_at': r[4],
                'category': r[5], 'priority': r[6],
            })
        return tickets

    return await asyncio.to_thread(fetch_tickets)


# --- ФУНКЦИИ АДМИНИСТРАТОРА И РЕЙТИНГА ---

async def get_admin_info(admin_id: int) -> dict | None:
    """Получает ФИО и средний рейтинг администратора."""

    def fetch_admin_info():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT full_name, total_rating, rating_count 
            FROM sys_admins WHERE telegram_id = ?
        """, (admin_id,))
        result = cursor.fetchone()

        conn.close()
        if result:
            full_name, total_rating, rating_count = result
            avg_rating = total_rating / rating_count if rating_count > 0 else 0.0
            return {
                'full_name': full_name, 'avg_rating': round(avg_rating, 2),
            }
        return None

    return await asyncio.to_thread(fetch_admin_info)


async def register_sys_admin(telegram_id: int, full_name: str, position: str) -> bool:
    """Регистрирует пользователя как SysAdmin в sys_admins и обновляет role в authorized_users."""

    def register():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        try:
            cursor.execute("UPDATE authorized_users SET role = 'admin' WHERE telegram_id = ?", (telegram_id,))

            cursor.execute("""
                INSERT OR REPLACE INTO sys_admins (telegram_id, full_name, rating_count, total_rating) 
                VALUES (?, ?, COALESCE((SELECT rating_count FROM sys_admins WHERE telegram_id = ?), 0), COALESCE((SELECT total_rating FROM sys_admins WHERE telegram_id = ?), 0))
            """, (telegram_id, full_name, telegram_id, telegram_id))

        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка регистрации SysAdmin {telegram_id}: {e}")
            return False
        finally:
            conn.commit()
            conn.close()
            return True

    return await asyncio.to_thread(register)


async def update_admin_rating(admin_id: int, rating: int):
    """Обновляет рейтинг администратора и average_rating."""

    def update_rating():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("SELECT total_rating, rating_count FROM sys_admins WHERE telegram_id = ?", (admin_id,))
        result = cursor.fetchone()

        if not result:
            logger.warning(f"DB: SysAdmin {admin_id} не найден для обновления рейтинга.")
            return

        total_rating, rating_count = result

        new_total_rating = total_rating + rating
        new_rating_count = rating_count + 1

        cursor.execute("""
            UPDATE sys_admins SET total_rating = ?, rating_count = ?
            WHERE telegram_id = ?
        """, (new_total_rating, new_rating_count, admin_id))

        conn.commit()
        conn.close()
        logger.info(
            f"DB: Рейтинг SysAdmin {admin_id} обновлен. Новая оценка: {rating}. Всего оценок: {new_rating_count}")

    await asyncio.to_thread(update_rating)


async def close_ticket_for_rating(ticket_id: int, admin_id: int) -> int | None:
    """Обновляет статус заявки на 'await_rating' и возвращает user_id создателя."""

    def close_ticket():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE tickets 
            SET status = 'await_rating', admin_id = ?, closed_at = CURRENT_TIMESTAMP
            WHERE id = ? AND status != 'await_rating'
        """, (admin_id, ticket_id))

        if cursor.rowcount == 0:
            return None

        cursor.execute("SELECT user_id FROM tickets WHERE id = ?", (ticket_id,))
        user_id = cursor.fetchone()[0]

        conn.commit()
        conn.close()
        return user_id

    return await asyncio.to_thread(close_ticket)


async def finalize_ticket_rating(ticket_id: int, rating: int) -> dict | None:
    """Записывает оценку, обновляет статус заявки и возвращает данные для обновления рейтинга."""

    def finalize():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("SELECT admin_id, ticket_number, user_id FROM tickets WHERE id = ?", (ticket_id,))
        result = cursor.fetchone()
        if not result: return None

        admin_id, ticket_number, user_id = result

        if not admin_id:
            logger.error(f"DB: Заявка {ticket_id} не была назначена администратору.")
            return None

        cursor.execute("UPDATE tickets SET status = 'closed' WHERE id = ?", (ticket_id,))

        conn.commit()
        conn.close()

        return {
            'admin_id': admin_id,
            'ticket_number': ticket_number,
            'user_id': user_id,
            'rating': rating
        }

    return await asyncio.to_thread(finalize)


# --- ФУНКЦИИ УПРАВЛЕНИЯ ОБОРУДОВАНИЕМ ---

async def create_equipment(inv_number: str, model: str, serial: str, category: str) -> bool:
    """Создает запись об оборудовании в таблице equipment."""

    def _create():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO equipment (inv_number, model, serial, category)
                VALUES (?, ?, ?, ?)
            """, (inv_number, model, serial, category))
            conn.commit()
            return True
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка создания оборудования {inv_number}: {e}")
            return False
        finally:
            conn.close()

    return await asyncio.to_thread(_create)


async def get_equipment(equipment_id: int = None, inv_number: str = None) -> Dict | None:
    """Получает информацию об оборудовании по ID или инвентарному номеру."""

    def fetch_equipment():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        if equipment_id:
            cursor.execute(
                "SELECT id, inv_number, model, serial, category, status, assigned_at, user_id FROM equipment WHERE id = ?",
                (equipment_id,))
        elif inv_number:
            cursor.execute(
                "SELECT id, inv_number, model, serial, category, status, assigned_at, user_id FROM equipment WHERE inv_number = ?",
                (inv_number,))
        else:
            return None

        r = cursor.fetchone()
        conn.close()
        if not r: return None
        return {
            'id': r[0], 'inv_number': r[1], 'model': r[2], 'serial': r[3], 'category': r[4],
            'status': r[5], 'assigned_at': r[6], 'user_id': r[7]
        }

    return await asyncio.to_thread(fetch_equipment)


async def assign_equipment_to_user(equipment_id: int, user_id: int, assigned_by: int) -> bool:
    """Назначает оборудование пользователю и создает запись в истории."""

    def _assign():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        try:
            cursor.execute(
                "UPDATE equipment SET user_id = ?, status = 'assigned', assigned_at = CURRENT_TIMESTAMP WHERE id = ?",
                (user_id, equipment_id))

            cursor.execute("""
                INSERT INTO equipment_history (equipment_id, from_user_id, to_user_id, assigned_by)
                VALUES (?, NULL, ?, ?)
            """, (equipment_id, user_id, assigned_by))

            conn.commit()
            return True
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка назначения оборудования {equipment_id} пользователю {user_id}: {e}")
            return False
        finally:
            conn.close()

    return await asyncio.to_thread(_assign)


async def get_all_equipment(status: str = None) -> List[Dict]:
    """Получает список всего оборудования с информацией о пользователях."""

    def fetch_all():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        query = """
            SELECT e.id, e.inv_number, e.model, e.serial, e.category, e.status, e.assigned_at, 
                   u.full_name
            FROM equipment e
            LEFT JOIN authorized_users u ON e.user_id = u.telegram_id
            WHERE 1=1
        """
        params = []
        if status: query += " AND e.status = ?"; params.append(status)
        query += " ORDER BY e.created_at DESC"

        cursor.execute(query, params)

        equipment = []
        for row in cursor.fetchall():
            equipment.append({
                'id': row[0], 'inv_number': row[1], 'model': row[2], 'serial': row[3], 'category': row[4],
                'status': row[5], 'assigned_at': row[6], 'user_name': row[7]
            })
        conn.close()
        return equipment

    return await asyncio.to_thread(fetch_all)


async def delete_equipment(inv_number: str) -> bool:
    """Удаляет оборудование по инвентарному номеру."""

    def _delete():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM equipment WHERE inv_number = ?", (inv_number,))
            conn.commit()
            return cursor.rowcount > 0
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка удаления оборудования {inv_number}: {e}")
            return False
        finally:
            conn.close()

    return await asyncio.to_thread(_delete)


# --- ФУНКЦИИ ДЛЯ ПОЛЬЗОВАТЕЛЬСКОГО ИНТЕРФЕЙСА (ПОВТОРНОЕ ОПРЕДЕЛЕНИЕ) ---

async def get_user_equipment(user_id: int) -> List[Dict]:
    """Получает список оборудования, назначенного пользователю."""

    def fetch_eq():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id, inv_number, model, serial, category, status, assigned_at
            FROM equipment
            WHERE user_id = ? AND status = 'assigned'
            ORDER BY assigned_at DESC
        """, (user_id,))

        equipment = []
        for row in cursor.fetchall():
            equipment.append({
                'id': row[0], 'inv_number': row[1], 'model': row[2], 'serial': row[3], 'category': row[4],
                'status': row[5], 'assigned_at': row[6]
            })
        conn.close()
        return equipment

    return await asyncio.to_thread(fetch_eq)


async def get_user_tickets(user_id: int) -> List[Dict]:
    """Получает список заявок пользователя."""

    def fetch_tickets():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, ticket_number, title, status, created_at, category, priority
            FROM tickets WHERE user_id = ? 
            ORDER BY created_at DESC
        """, (user_id,))

        rows = cursor.fetchall()
        conn.close()

        tickets = []
        for r in rows:
            tickets.append({
                'id': r[0], 'number': r[1], 'title': r[2], 'status': r[3], 'created_at': r[4],
                'category': r[5], 'priority': r[6],
            })
        return tickets

    return await asyncio.to_thread(fetch_tickets)


# --- ФУНКЦИИ УПРАВЛЕНИЯ РАБОЧИМИ МЕСТАМИ ---

async def get_workplace(number: str) -> Dict | None:
    """Алиас для get_workplace_by_number (Устраняет ошибку импорта в handlers/workplaces.py)."""
    return await get_workplace_by_number(number)


async def get_workplace_by_number(number: str) -> Dict | None:
    """Возвращает рабочее место по его номеру (number)."""

    def _get():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, number, department, location, floor, primary_pc, peripherals FROM workplaces WHERE number = ?",
            (number,))
        r = cursor.fetchone()
        conn.close()
        if not r: return None
        return {'id': r[0], 'number': r[1], 'department': r[2], 'location': r[3], 'floor': r[4], 'primary_pc': r[5],
                'peripherals': r[6]}

    return await asyncio.to_thread(_get)


async def get_workplace_equipment(workplace_id: int) -> List[Dict]:
    """Получает список оборудования, закрепленного за рабочим местом."""

    def fetch_eq():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id, inv_number, model, serial, category, status, user_id
            FROM equipment
            WHERE workplace_id = ?
            ORDER BY inv_number
        """, (workplace_id,))

        equipment = []
        for row in cursor.fetchall():
            equipment.append({
                'id': row[0], 'inv_number': row[1], 'model': row[2], 'serial': row[3], 'category': row[4],
                'status': row[5], 'user_id': row[6]
            })
        conn.close()
        return equipment

    return await asyncio.to_thread(fetch_eq)


async def get_all_workplaces() -> List[Dict]:
    """Получает список всех рабочих мест."""

    def fetch_all():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id, number, department, location
            FROM workplaces
            ORDER BY number
        """)

        workplaces = []
        for row in cursor.fetchall():
            workplaces.append({
                'id': row[0],
                'number': row[1],
                'department': row[2],
                'location': row[3]
            })
        conn.close()
        return workplaces

    return await asyncio.to_thread(fetch_all)


async def create_workplace(number: str, department: str, location: str) -> bool:
    """Создает новое рабочее место с указанными параметрами."""

    def _create():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO workplaces (number, department, location)
                VALUES (?, ?, ?)
            """, (number, department, location))
            conn.commit()
            return True
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка создания рабочего места {number}: {e}")
            return False
        finally:
            conn.close()

    return await asyncio.to_thread(_create)


async def delete_workplace(number: str) -> bool:
    """Удаляет рабочее место по его номеру."""

    def _delete():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM workplaces WHERE number = ?", (number,))
            conn.commit()
            return cursor.rowcount > 0
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка удаления рабочего места {number}: {e}")
            return False
        finally:
            conn.close()

    return await asyncio.to_thread(_delete)


# --- МЕТОДЫ ОПРЕДЕЛЕНИЯ ЭТАЖА И ДРУГИЕ УТИЛИТЫ ---

KNOWN_IP_FLOOR: Dict[str, int] = {
    '172.20.30.107': 2, '172.20.30.110': 2, '172.20.30.132': 2,
    '172.20.30.36': 4, '172.20.30.48': 4,
    '172.20.31.1': 5, '172.20.31.27': 5,
}


def get_floor_from_hostname(hostname: str) -> Optional[int]:
    """Определяет этаж по имени компьютера (TSS-WS-5xxx -> 5 этаж)."""
    if not hostname: return None
    name = str(hostname).upper().strip()
    m = re.match(r'^TSS-WS-(\d+)', name)
    if not m: return None
    num = m.group(1)
    if num.startswith('5'): return 5
    if num.startswith('2'): return 2
    return 4


def get_floor_from_ip(ip: str, hostname: Optional[str] = None) -> Optional[int]:
    """Определяет этаж по IP, с fallback на hostname."""
    if not ip: return get_floor_from_hostname(hostname) if hostname else None
    ip = ip.strip()
    if ip in KNOWN_IP_FLOOR: return KNOWN_IP_FLOOR[ip]
    if ip.startswith('172.20.31.'): return 5
    if hostname: return get_floor_from_hostname(hostname)
    return None


async def get_available_floors() -> List[int]:
    """Возвращает список уникальных номеров этажей из таблицы workplaces."""

    def fetch_floors():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT floor FROM workplaces WHERE floor IS NOT NULL ORDER BY floor ASC")
        floors = [row[0] for row in cursor.fetchall()]
        conn.close()
        allowed = {2, 4, 5}
        filtered = [f for f in floors if f in allowed]
        return filtered if filtered else sorted(list(allowed))

    return await asyncio.to_thread(fetch_floors)


async def get_workplaces_by_floor(floor: int) -> List[Dict]:
    """Возвращает рабочие места для указанного этажа."""

    def fetch_workplaces():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT number, primary_pc FROM workplaces WHERE floor = ? ORDER BY number ASC", (floor,))
        workplaces = []
        for row in cursor.fetchall():
            workplaces.append({'number': row[0], 'pc_name': row[1]})
        conn.close()
        return workplaces

    workplaces = await asyncio.to_thread(fetch_workplaces)
    # Если таблица пустая (новая БД) — повторно выполним автосоздание и перечитаем.
    if not workplaces:
        await asyncio.to_thread(_seed_default_workplaces)
        workplaces = await asyncio.to_thread(fetch_workplaces)
    return workplaces


async def get_all_users_for_mailing() -> list[int]:
    """Получает Telegram ID всех активных пользователей для рассылки."""

    def _get_ids():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT telegram_id FROM authorized_users")
        user_ids = [row[0] for row in cursor.fetchall()]
        conn.close()
        return user_ids

    return await asyncio.to_thread(_get_ids)


# --- Вложения к заявкам ---

async def add_ticket_attachment(ticket_id: int, file_id: str, file_type: str = "photo",
                                file_name: Optional[str] = None) -> bool:
    """Сохранить вложение к заявке."""

    def _add():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO ticket_attachments (ticket_id, file_id, file_type, file_name)
                VALUES (?, ?, ?, ?)
            """, (ticket_id, file_id, file_type, file_name))
            conn.commit()
            return True
        except sqlite3.Error as e:
            logger.error(f"DB: ошибка добавления вложения для заявки {ticket_id}: {e}")
            return False
        finally:
            conn.close()

    return await asyncio.to_thread(_add)


# --- FAQ ---

async def add_faq_item(question: str, answer: str) -> bool:
    """Добавить пункт в FAQ."""

    def _add():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO faq (question, answer) VALUES (?, ?)", (question, answer))
            conn.commit()
            return True
        except sqlite3.Error as e:
            logger.error(f"DB: ошибка добавления FAQ: {e}")
            return False
        finally:
            conn.close()

    return await asyncio.to_thread(_add)


async def get_faq_items(limit: Optional[int] = None) -> List[Dict]:
    """Получить список FAQ (последние сверху)."""

    def _get():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        query = "SELECT id, question, answer, created_at FROM faq ORDER BY created_at DESC"
        if limit:
            query += f" LIMIT {int(limit)}"
        cursor.execute(query)
        items = []
        for row in cursor.fetchall():
            items.append({
                "id": row[0],
                "question": row[1],
                "answer": row[2],
                "created_at": row[3],
            })
        conn.close()
        return items

    return await asyncio.to_thread(_get)


async def delete_faq_item(faq_id: int) -> bool:
    """Удалить пункт FAQ."""

    def _delete():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM faq WHERE id = ?", (faq_id,))
            conn.commit()
            return cursor.rowcount > 0
        except sqlite3.Error as e:
            logger.error(f"DB: ошибка удаления FAQ {faq_id}: {e}")
            return False
        finally:
            conn.close()

    return await asyncio.to_thread(_delete)


async def save_faq_material(title: str, description: str, file_info: Optional[Dict] = None) -> Dict | None:
    """Сохраняет новый материал FAQ и возвращает его данные для рассылки."""

    def _save():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        file_id = file_info.get('id') if file_info else None
        file_type = file_info.get('type') if file_info else None

        try:
            # Убедитесь, что таблица 'faq_materials' создана в init_db
            cursor.execute("""
                INSERT INTO faq_materials (title, description, file_id, file_type)
                VALUES (?, ?, ?, ?)
            """, (title, description, file_id, file_type))

            faq_id = cursor.lastrowid
            conn.commit()
            return {'id': faq_id, 'title': title, 'description': description, 'file_id': file_id,
                    'file_type': file_type}
        except sqlite3.Error as e:
            logger.error(f"DB: Ошибка сохранения FAQ: {e}")
            return None
        finally:
            conn.close()

    return await asyncio.to_thread(_save)


async def get_latest_faq_material() -> Dict | None:
    """Получает самый свежий материал FAQ для рассылки."""

    def _get():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, title, description, file_id, file_type 
            FROM faq_materials 
            ORDER BY created_at DESC LIMIT 1
        """)
        row = cursor.fetchone()
        conn.close()
        if row:
            return {
                'id': row[0], 'title': row[1], 'description': row[2], 'file_id': row[3], 'file_type': row[4],
            }
        return None

    return await asyncio.to_thread(_get)


async def get_all_faq_materials() -> list[dict]:
    """Получает список всех сохраненных материалов FAQ/гайдов."""

    def _get_all():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, title, description, file_id, file_type 
            FROM faq_materials 
            ORDER BY created_at DESC
        """)
        columns = [desc[0] for desc in cursor.description]
        # Преобразуем результат в список словарей для удобства
        faq_list = [dict(zip(columns, row)) for row in cursor.fetchall()]
        conn.close()
        return faq_list

    return await asyncio.to_thread(_get_all)


async def get_user_credentials(telegram_id: int) -> List[Dict]:
    """Получить сохраненные доступы пользователя."""

    def _get():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT service, login, password, url, note, created_at
            FROM user_credentials
            WHERE telegram_id = ?
            ORDER BY created_at DESC
        """, (telegram_id,))
        rows = cursor.fetchall()
        conn.close()
        creds = []
        for r in rows:
            creds.append({
                "service": r[0],
                "login": r[1],
                "password": r[2],
                "url": r[3],
                "note": r[4],
                "created_at": r[5],
            })
        return creds

    return await asyncio.to_thread(_get)


